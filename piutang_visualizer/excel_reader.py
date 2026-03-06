"""
Module untuk membaca dan memproses data dari file Excel
"""

import glob
import os
import warnings
from datetime import datetime
from typing import List, Dict, Tuple, Optional, Any
from openpyxl import load_workbook

from .config import COLUMN_MAPPING

# Suppress openpyxl UserWarning
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class ExcelReader:
    """Kelas untuk membaca file Excel"""
    
    def __init__(self, filepath: Optional[str] = None):
        self.filepath = filepath
        self.headers: List[str] = []
        self.raw_data: List[List[str]] = []
        self.processed_data: List[Dict[str, Any]] = []
    
    @staticmethod
    def find_excel_files(directory: str = ".") -> List[str]:
        """Cari semua file .xlsx di direktori (kecuali file temp)"""
        pattern = os.path.join(directory, "*.xlsx")
        files = glob.glob(pattern)
        # Filter out temporary files (yang dimulai dengan ~$)
        return [f for f in files if not os.path.basename(f).startswith("~$")]
    
    def load(self, filepath: Optional[str] = None) -> 'ExcelReader':
        """Load data dari file Excel"""
        if filepath:
            self.filepath = filepath
        
        if not self.filepath:
            raise ValueError("Filepath tidak ditemukan")
        
        wb = load_workbook(self.filepath, data_only=True)
        ws = wb.active
        
        self.headers = []
        self.raw_data = []
        
        for i, row in enumerate(ws.iter_rows()):
            row_values = [
                str(cell.value) if cell.value is not None else "" 
                for cell in row
            ]
            
            # Skip empty rows
            if not any(row_values):
                continue
            
            if i == 0:
                self.headers = row_values
            else:
                self.raw_data.append(row_values)
        
        return self
    
    def _find_column_index(self, column_type: str) -> Optional[int]:
        """Cari index kolom berdasarkan tipe"""
        keywords = COLUMN_MAPPING.get(column_type, [])
        
        # Prioritaskan exact match dulu
        for i, header in enumerate(self.headers):
            if not header or header == 'None':
                continue
            header_clean = header.lower().strip()
            for keyword in keywords:
                # Exact match atau match diawal string (hindari 'kategori' match 'nama')
                if keyword == header_clean or header_clean.startswith(keyword):
                    return i
        
        # Fallback: substring match tapi lebih ketat
        for i, header in enumerate(self.headers):
            if not header or header == 'None':
                continue
            header_clean = header.lower().strip()
            for keyword in keywords:
                if keyword in header_clean:
                    # Skip kalau ada kata 'kategori' tapi kita cari 'nama pelanggan'
                    if column_type == 'nama_pelanggan' and 'kategori' in header_clean:
                        continue
                    return i
        return None
    
    def _parse_number(self, value: str) -> float:
        """Parse string ke float dengan membersihkan format
        
        Nilai negatif atau dalam kurung (xxx) dianggap kredit/lebih bayar,
        return 0 karena bukan piutang.
        """
        if not value or value == 'None':
            return 0.0
        
        val_str = str(value).strip()
        
        # Cek nilai negatif: diawali - atau diapit ()
        # NOTE: Cek dulu SEBELUM hapus karakter - atau ()
        is_negative = val_str.startswith('-') or (val_str.startswith('(') and val_str.endswith(')'))
        
        # Bersihkan format: hapus (), -, Rp, spasi
        cleaned = val_str.replace('(', '').replace(')', '').replace('-', '').replace('Rp', '').replace('rp', '').replace(' ', '')
        
        # Handle format angka: 
        # - Format US: 1,234.56 -> hapus koma, titik desimal
        # - Format ID: 1.234,56 -> hapus titik, ganti koma jadi titik
        if ',' in cleaned and '.' in cleaned:
            # Ada koma dan titik, tentukan format
            last_comma = cleaned.rfind(',')
            last_dot = cleaned.rfind('.')
            if last_comma > last_dot:
                # Format ID: 1.234,56 -> hapus titik, ganti koma jadi titik
                cleaned = cleaned.replace('.', '').replace(',', '.')
            else:
                # Format US: 1,234.56 -> hapus koma
                cleaned = cleaned.replace(',', '')
        elif ',' in cleaned:
            # Cek apakah koma sebagai desimal atau ribuan
            # Jika ada 3 digit setelah koma -> ribuan, hapus koma
            # Jika tidak -> desimal, ganti jadi titik
            parts = cleaned.split(',')
            if len(parts) == 2 and len(parts[1]) == 3:
                # Ribuan: 199,428 -> 199428
                cleaned = cleaned.replace(',', '')
            else:
                # Desimal: 199,75 -> 199.75
                cleaned = cleaned.replace(',', '.')
        
        try:
            num = float(cleaned) if cleaned else 0.0
            # Jika negatif, return 0 (bukan piutang)
            return 0.0 if is_negative else num
        except ValueError:
            return 0.0
    
    def _parse_integer(self, value: str) -> int:
        """Parse string ke integer"""
        return int(self._parse_number(value))
    
    def _format_date(self, date_value: str) -> str:
        """Format tanggal ke dd/mm/yyyy"""
        if not date_value or date_value in ['None', '']:
            return "-"
        
        try:
            # Coba parse jika format datetime
            if ' ' in str(date_value):
                dt = datetime.strptime(str(date_value).split()[0], '%Y-%m-%d')
                return dt.strftime('%d/%m/%Y')
            # Jika sudah format date
            elif '-' in str(date_value) and len(str(date_value)) == 10:
                dt = datetime.strptime(str(date_value), '%Y-%m-%d')
                return dt.strftime('%d/%m/%Y')
            return date_value
        except:
            return date_value
    
    def _format_rupiah(self, amount: float) -> str:
        """Format angka ke Rupiah"""
        return f"Rp {amount:,.0f}".replace(',', '.')
    
    def process(self) -> List[Dict[str, Any]]:
        """Proses raw data ke structured data"""
        # Cari index kolom
        col_indices = {
            'nama_pelanggan': self._find_column_index('nama_pelanggan'),
            'no_faktur': self._find_column_index('no_faktur'),
            'tgl_faktur': self._find_column_index('tgl_faktur'),
            'piutang': self._find_column_index('piutang'),
            'umur': self._find_column_index('umur')
        }
        
        self.processed_data = []
        
        for row in self.raw_data:
            try:
                item = {
                    'nama_pelanggan': (
                        row[col_indices['nama_pelanggan']] 
                        if col_indices['nama_pelanggan'] is not None else ""
                    ),
                    'no_faktur': (
                        row[col_indices['no_faktur']] 
                        if col_indices['no_faktur'] is not None else ""
                    ),
                    'tgl_faktur': (
                        row[col_indices['tgl_faktur']] 
                        if col_indices['tgl_faktur'] is not None else ""
                    ),
                    'piutang_raw': (
                        row[col_indices['piutang']] 
                        if col_indices['piutang'] is not None else "0"
                    ),
                    'umur_raw': (
                        row[col_indices['umur']] 
                        if col_indices['umur'] is not None else "0"
                    )
                }
                
                # Parse dan format data
                item['piutang'] = self._parse_number(item['piutang_raw'])
                item['umur'] = self._parse_integer(item['umur_raw'])
                
                # Skip jika piutang <= 0 (bukan piutang, tapi kredit/lebih bayar)
                if item['piutang'] <= 0:
                    continue
                
                item['piutang_fmt'] = self._format_rupiah(item['piutang'])
                item['tgl_faktur_fmt'] = self._format_date(item['tgl_faktur'])
                
                self.processed_data.append(item)
                
            except Exception as e:
                # Skip row yang error
                continue
        
        return self.processed_data
    
    def get_unique_customers(self) -> List[str]:
        """Dapatkan daftar nama pelanggan unik"""
        customers = set()
        for item in self.processed_data:
            name = item.get('nama_pelanggan', '')
            if name and name != 'None':
                customers.add(name)
        return sorted(list(customers))
    
    def filter_by_customer(self, customer_name: str) -> List[Dict[str, Any]]:
        """Filter data berdasarkan nama pelanggan"""
        return [
            item for item in self.processed_data 
            if item.get('nama_pelanggan') == customer_name
        ]
    
    def get_stats(self) -> Dict[str, Any]:
        """Dapatkan statistik data"""
        if not self.processed_data:
            return {}
        
        total_piutang = sum(item['piutang'] for item in self.processed_data)
        
        return {
            'total_rows': len(self.processed_data),
            'unique_customers': len(self.get_unique_customers()),
            'total_piutang': total_piutang
        }


def load_excel_data(filepath: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Helper function untuk load Excel (backward compatible)"""
    reader = ExcelReader(filepath)
    reader.load()
    data = reader.process()
    return reader.headers, data
